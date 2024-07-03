import json
import openpyxl

# Načtení JSON souborů
try:
    with open('skoly.json', 'r', encoding='utf-8') as file:
        data = json.load(file)
        print("Data loaded successfully.")
except Exception as e:
    print(f"Error loading data.json: {e}")

try:
    with open('kraje.json', 'r', encoding='utf-8') as file:
        kraje_data = json.load(file)
        print("Kraje loaded successfully.")
except Exception as e:
    print(f"Error loading kraje.json: {e}")

# Vytvoření mapy ID krajů na názvy krajů
kraje = {kraj['id']: kraj['nazev']['cs'] for kraj in kraje_data['polozky']}
print(f"Kraje map: {kraje}")

# Funkce pro získání názvu kraje podle ID
def get_kraj_name(kraj_id):
    return kraje.get(kraj_id, 'Neznámý kraj')

# Roztřídění škol podle krajů
schools_by_kraj = {}

for item in data['polozky']:
    kraj_id = item['adresaSidla']['kraj']['id']
    kraj_name = get_kraj_name(kraj_id)

    if kraj_name not in schools_by_kraj:
        schools_by_kraj[kraj_name] = []
    
    school_info = {
        'nazev': item['nazev'],
        'email': item['email'],
        'urlAdresa': item['urlAdresa'],
        'reditel': item['reditel'],
        'reditelTelefon': item['reditelTelefon'],
        'kontaktniOsoba': item['kontaktniOsoba'],
        'kontaktniOsobaTelefon': item['kontaktniOsobaTelefon']
    }
    
    schools_by_kraj[kraj_name].append(school_info)

print("Schools sorted by regions successfully.")

# Vytvoření Excel souboru a zápis dat
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Školy podle krajů"

# Zápis záhlaví
headers = ["Název", "Email", "urlAdresa", "Ředitel", "Telefon ředitele", "kontaktniOsoba", "kontaktniOsobaTelefon"]

# Iterace přes kraje a zápis dat do samostatných listů
for kraj, schools in schools_by_kraj.items():
    ws = wb.create_sheet(title=kraj[:31])  # List může mít maximálně 31 znaků v názvu
    ws.append(headers)
    
    for school in schools:
        ws.append([school['nazev'], school['email'], school['urlAdresa'], school['reditel'], school['reditelTelefon'], school['kontaktniOsoba'], school['kontaktniOsobaTelefon']])

# Odebrání výchozího listu, pokud je prázdný
if "Sheet" in wb.sheetnames:
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

# Uložení Excel souboru
output_file = "Skoly_podle_kraju.xlsx"
wb.save(output_file)
print(f"Data successfully written to {output_file}")